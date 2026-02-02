import os
import csv
import tempfile
import re
from typing import List, Dict, Optional
import openpyxl
from io import BytesIO
from datetime import date
import pandas as pd
from backend.core.config import DATA_DIR
from backend.models.excel import ExcelFile, ExcelCell
from backend.models.dataset import Dataset, DatasetStatus
from backend.core.database import Base
from backend.services.dataset_service import _normalize_header


FILE_DEFINITIONS = {
    "schedule_input": {
        "display_name": "勤務予定入力",
        "expected_headers": [
            "従業員番号",
            "勤務予定日",
            "出勤休日区分",
            "出勤休日区分名",
            "就業時間パターンコード",
            "就業時間パターン名",
            "就業開始時刻",
            "就業終了時刻",
            "休憩時間",
        ],
    },
    "punches": {
        "display_name": "出退社時刻",
        "expected_headers": [
            "従業員番号",
            "勤務日付",
            "出社時刻",
            "退社時刻",
        ],
    },
    "days_items": {
        "display_name": "日数項目",
        "expected_headers": [
            "従業員番号",
            "勤務日",
            "出社時刻",
            "退社時刻",
            "日数項目",
            "日数項目名",
        ],
    },
    "tim_daily": {
        "display_name": "日次実績",
        "expected_headers": [
            "従業員番号",
            "勤務日付",
            "(時間)定時開始時刻",
            "(時間)定時終了時刻",
            "(時間)呼出出勤",
            "(時間)呼出退勤",
            "(時間)呼出勤務",
            "(時間)実所定外時間",
            "(時間)出社日数",
            "(時間)在宅勤務時間",
            "(時間)在宅勤務日数",
            "(時間)終日在宅フラグ",
            "(時間)実労働時間",
            "(時間)休憩Ｈ",
            "(時間)休憩勤務開始",
            "(時間)休憩勤務終了",
            "(時間)休憩1開始時刻",
            "(時間)休憩1終了時刻",
            "(時間)休憩2開始時刻",
            "(時間)休憩2終了時刻",
            "(時間)休憩3開始時刻",
            "(時間)休憩3終了時刻",
            "(時間)休憩4開始時刻",
            "(時間)休憩4終了時刻",
        ],
    },
    "person_progress": {
        "display_name": "勤務予定進捗一覧",
        "expected_headers": [
            "社員番号",
            "氏名",
            "カナ氏名",
            "勤怠年月",
            "勤務開始日",
            "進捗状況",
            "打刻実績",
            "勤務実績登録",
            "所属名称",
            "メールアドレス",
        ],
    },
    "org_info": {
        "display_name": "所属情報",
        "expected_headers": [
            "(基本)従業員番号",
            "(基本)氏名",
            "(人事所属本務(基準日))所属コード",
            "(人事所属本務(基準日))所属名称１",
            "(人事所属本務(基準日))所属名称２",
            "(人事所属本務(基準日))所属名称３",
            "(人事所属本務(基準日))所属名称４",
            "(人事所属本務(基準日))所属名称５",
            "(人事所属本務(基準日))所属名称６",
            "(人事所属本務(基準日))所属名称７",
            "(人事所属本務(基準日))所属名称８",
            "(従業員区分(基準日))従業員区分(ｺｰﾄﾞ)",
            "(従業員区分(基準日))従業員区分",
            "(職制(基準日))職制(ｺｰﾄﾞ)",
            "(職制(基準日))職制",
            "(人事所属本務(基準日))損益管理コード(ｺｰﾄﾞ)",
            "(人事所属本務(基準日))損益管理コード",
            "(メールアドレス情報)アドレス1",
            "(基本)入社年月日",
        ],
    },
}


def normalize_headers(headers: List[str]) -> List[str]:
    cleaned = []
    for h in headers:
        if h is None:
            h = ""
        if not isinstance(h, str):
            h = str(h)
        cleaned.append(h.lstrip("\ufeff").strip())
    return cleaned


def parse_csv_to_cells(path: str):
    cells = []
    headers: List[str] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, start=1):
            if r_idx == 1:
                headers = normalize_headers(row)
            for c_idx, val in enumerate(row, start=1):
                cells.append((r_idx, c_idx, val))
    return cells, headers, "Sheet1"


def parse_xlsx_to_cells(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    cells = []
    headers: List[str] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            headers = normalize_headers(list(row))
        for c_idx, val in enumerate(row, start=1):
            cells.append((r_idx, c_idx, "" if val is None else str(val)))
    return cells, headers, ws.title or "Sheet1"


def fetch_grid_for_key(db, file_key: str):
    dataset = (
        db.query(Dataset)
        .filter(Dataset.kind == file_key, Dataset.status == DatasetStatus.ready)
        .order_by(Dataset.uploaded_at.desc())
        .first()
    )
    if dataset:
        try:
            df = pd.read_parquet(dataset.stored_path)
            headers = list(df.columns)
            rows = df.fillna("").astype(str).values.tolist()
            grid = [headers, *rows]
            return {"grid": grid, "headers": headers, "rows": rows}
        except Exception:
            pass

    # fallback legacy path
    ef = (
        db.query(ExcelFile)
        .filter(ExcelFile.file_key == file_key)
        .order_by(ExcelFile.version.desc())
        .first()
    )
    if not ef:
        return None
    cells = (
        db.query(ExcelCell)
        .filter(ExcelCell.file_id == ef.id)
        .order_by(ExcelCell.sheet_name, ExcelCell.row_index, ExcelCell.col_index)
        .all()
    )
    sheets = {}
    for cell in cells:
        sheets.setdefault(cell.sheet_name, []).append(cell)
    if not sheets:
        return None
    sheet_name, sheet_cells = list(sheets.items())[0]
    max_row = max(c.row_index for c in sheet_cells)
    max_col = max(c.col_index for c in sheet_cells)
    grid = [["" for _ in range(max_col)] for _ in range(max_row)]
    for c in sheet_cells:
        grid[c.row_index - 1][c.col_index - 1] = c.value or ""
    return {"grid": grid, "headers": grid[0] if grid else [], "rows": grid[1:] if len(grid) > 1 else []}


def _normalized_header_key(header: str) -> str:
    if header is None:
        return ""
    if not isinstance(header, str):
        header = str(header)
    h = header.lstrip("\ufeff").strip()
    h = re.sub(r"[()（）\[\]【】]", "", h)
    h = re.sub(r"^時間", "", h)
    h = h.replace(" ", "").replace("　", "").replace("/", "")
    return h.lower()


def _build_column_map(headers: List[str]) -> Dict[str, int]:
    normalized = {_normalized_header_key(h): idx for idx, h in enumerate(headers)}
    aliases = {
        "emp_no": ["従業員番号", "社員番号", "社員No"],
        "name": ["氏名", "名前", "カナ氏名"],
        "status": ["勤務予定", "勤務予定日", "勤務予定区分", "勤務状況", "進捗状況"],
        "overtime": ["実所定外時間", "残業時間", "残業", "(時間)実所定外時間"],
        "overtime_detail": ["残業時間", "実所定外時間", "(時間)残業時間"],
        "call_time": ["呼出出勤時間", "呼出出勤", "(時間)呼出出勤"],
        "grade": [
            "グレード",
            "従業員区分",
            "キャリアグレード",
            "キャリア グレード",
            "所属情報のキャリアグレード",
            "(従業員区分(基準日))従業員区分",
        ],
        "role": ["職制", "役職"],
        "org2": ["所属名称2", "所属名称２", "所属2", "所属２"],
        "org3": ["所属名称3", "所属名称３", "所属3", "所属３"],
        "org4": ["所属名称4", "所属名称４", "所属4", "所属４"],
        "org5": ["所属名称5", "所属名称５", "所属5", "所属５"],
        "org6": ["所属名称6", "所属名称６", "所属6", "所属６", "所属情報6", "所属情報６"],
        "org7": ["所属名称7", "所属名称７", "所属7", "所属７"],
        "org8": ["所属名称8", "所属名称８", "所属8", "所属８"],
    }

    resolved = {}
    for key, candidates in aliases.items():
        for c in candidates:
            idx = normalized.get(_normalized_header_key(c))
            if idx is not None:
                resolved[key] = idx
                break
    return resolved


def _build_rows_from_grid(grid: Dict) -> List[List[str]]:
    headers = grid.get("headers") or []
    rows = grid.get("rows") or []
    col_map = _build_column_map(headers)

    if "emp_no" not in col_map:
        raise ValueError("必須列「従業員番号」が見つかりませんでした。")

    data_rows: List[List[str]] = []
    for row in rows:
        def pick(key: str, default: str = "") -> str:
            idx = col_map.get(key)
            if idx is None:
                return default
            return row[idx] if idx < len(row) and row[idx] is not None else default

        overtime_value = pick("overtime", "0:00") or "0:00"

        # 所属名称: それぞれの列をそのまま出力（移動しない）
        org2 = pick("org2", "")
        org3 = pick("org3", "")
        org4 = pick("org4", "")
        org5 = pick("org5", "")
        org6 = pick("org6", "")
        org7 = pick("org7", "")
        org8 = pick("org8", "")

        data_rows.append(
            [
                pick("emp_no", ""),
                pick("name", ""),
                pick("status", ""),
                overtime_value,
                pick("overtime_detail", overtime_value),
                pick("call_time", "0:00"),
                pick("grade", ""),
                pick("role", ""),
                org2,
                org3,
                org4,
                org5,
                org6,
                org7,
                org8,
            ]
        )

    if not data_rows:
        data_rows.append([""] * 15)
    return data_rows


def build_processed_excel(grid: Dict, target_ym: str = "") -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "実所定外時間 推計データ"

    ym_label = target_ym or "2025年12月度"
    today_label = date.today().strftime("%Y年%m月%d日")
    title = f"{ym_label} 実所定外時間 推計データ（{today_label}現在）"
    ws.merge_cells("A1:O1")
    ws["A1"] = title
    ws["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    legends = [
        ("80h超", "長時間労働", "6b4f00", "f7f2e2"),
        ("〜80h", "３６協定特別条項上限超過者", "d0a754", "1a1200"),
        ("〜60h", "３６協定特別条項上限", "e6a600", "1a1200"),
        ("〜45h", "労働基準法上の時間外労働上限", "c7b202", "0f0f0f"),
        ("〜30h", "社内ルールに基づく上限", "1f8a55", "fdfdfd"),
        ("15h〜20h", "", "5f86c6", "fdfdfd"),
    ]
    start_row = 3
    for idx, (label, desc, bg, fg) in enumerate(legends):
        r = start_row + idx
        ws[f"A{r}"] = label
        ws[f"A{r}"].fill = openpyxl.styles.PatternFill("solid", fgColor=bg)
        ws[f"A{r}"].font = openpyxl.styles.Font(color=fg, bold=True)
        ws[f"A{r}"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws[f"B{r}"] = desc
        ws[f"B{r}"].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

    header_titles = [
        "従業員番号",
        "氏名",
        "勤務予定",
        "実所定外時間",
        "残業時間",
        "呼出出勤時間",
        "グレード",
        "職制",
        "所属名称2",
        "所属名称3",
        "所属名称4",
        "所属名称5",
        "所属名称6",
        "所属名称7",
        "所属名称8",
    ]
    header_row = start_row + len(legends) + 2
    sub_header_row = header_row + 1

    # 上段ヘッダ
    merges = {
        "A": "A",
        "B": "B",
        "C": "C",
        "D": "D",
        "E": "F",  # 内訳のグルーピング
        "G": "G",
        "H": "H",
        "I": "I",
        "J": "J",
        "K": "K",
        "L": "L",
        "M": "M",
        "N": "N",
        "O": "O",
    }
    for start_col, end_col in merges.items():
        if start_col != end_col:
            ws.merge_cells(f"{start_col}{header_row}:{end_col}{header_row}")

    header_labels_top = [
        "従業員番号",
        "氏名",
        "勤務予定",
        "実所定外時間",
        "内訳",
        None,
        "グレード",
        "職制",
        "所属名称2",
        "所属名称3",
        "所属名称4",
        "所属名称5",
        "所属名称6",
        "所属名称7",
        "所属名称8",
    ]
    header_labels_bottom = [
        "従業員番号",
        "氏名",
        "勤務予定",
        "実所定外時間",
        "残業時間",
        "呼出出勤時間",
        "グレード",
        "職制",
        "所属名称2",
        "所属名称3",
        "所属名称4",
        "所属名称5",
        "所属名称6",
        "所属名称7",
        "所属名称8",
    ]

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="ffffff")
    yellow_fill = openpyxl.styles.PatternFill("solid", fgColor="fff8a8")
    orange_fill = openpyxl.styles.PatternFill("solid", fgColor="f9d3b0")
    red_font = openpyxl.styles.Font(color="c81e1e", bold=True)
    default_font = openpyxl.styles.Font(bold=True)
    border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin", color="b7b7b7"),
        right=openpyxl.styles.Side(style="thin", color="b7b7b7"),
        top=openpyxl.styles.Side(style="thin", color="b7b7b7"),
        bottom=openpyxl.styles.Side(style="thin", color="b7b7b7"),
    )

    for idx, label in enumerate(header_labels_top, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.fill = header_fill
        cell.font = default_font
        if idx == 4:
            cell.fill = yellow_fill
        if idx == 7:
            cell.font = red_font
        if 9 <= idx <= 13:
            cell.fill = orange_fill
            cell.font = red_font
        if idx in (14, 15):
            cell.font = red_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        cell.border = border

    for idx, label in enumerate(header_labels_bottom, start=1):
        cell = ws.cell(row=sub_header_row, column=idx, value=label)
        cell.fill = header_fill
        cell.font = default_font
        if idx == 4:
            cell.fill = yellow_fill
        if idx == 7:
            cell.font = red_font
        if 9 <= idx <= 13:
            cell.fill = orange_fill
            cell.font = red_font
        if idx in (14, 15):
            cell.font = red_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[header_row].height = 24
    ws.row_dimensions[sub_header_row].height = 24

    try:
        data_rows = _build_rows_from_grid(grid)
    except ValueError as exc:
        raise

    data_start = sub_header_row + 1
    for r_idx, r in enumerate(data_rows, start=data_start):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            align = "center" if 4 <= c_idx <= 6 else "left"
            cell.alignment = openpyxl.styles.Alignment(horizontal=align, vertical="center")
            cell.border = border

    widths = [13, 14, 12, 13, 12, 14, 10, 10, 13, 13, 13, 13, 13, 13, 13]
    for c_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    ws.freeze_panes = ws["A" + str(data_start)]

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
