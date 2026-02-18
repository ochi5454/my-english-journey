"""宛先管理API（ファジー検索機能付き、AIエージェント対応、Entraバリデーション対応）"""
import io
import json
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
import pandas as pd
import httpx
from rapidfuzz import fuzz
import openai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from backend.core.database import get_db
from backend.core.auth import get_current_user, get_user_tokens
from backend.core.config import get_settings
from backend.models.user import User
from backend.models.recipient import RecipientList, RecipientListMember
from backend.services.entra_validation_service import EntraValidationService
from backend.services.langchain_recipient_filter import (
    LangChainRecipientFilter,
    MemberData,
    FilterResult as LangChainFilterResult,
)

router = APIRouter(prefix="/recipients", tags=["recipients"])
settings = get_settings()


# Pydantic Schemas
class RecipientMemberCreate(BaseModel):
    email: str
    name: Optional[str] = None
    department: Optional[str] = None
    position: Optional[str] = None  # 役職（チーフ、マネージャー等）
    employee_id: Optional[str] = None  # 社員番号
    note: Optional[str] = None
    recipient_type: Optional[str] = 'to'  # 'to', 'cc', 'bcc'


class RecipientListCreate(BaseModel):
    name: str
    description: Optional[str] = None
    members: Optional[List[RecipientMemberCreate]] = None


class RecipientMemberResponse(BaseModel):
    id: int
    email: str
    name: Optional[str]
    department: Optional[str]
    position: Optional[str]  # 役職
    employee_id: Optional[str] = None  # 社員番号
    note: Optional[str]
    recipient_type: str = 'to'  # 'to', 'cc', 'bcc'

    class Config:
        from_attributes = True


class RecipientListResponse(BaseModel):
    id: int
    name: str
    description: Optional[str]
    member_count: int
    to_count: int = 0  # To宛先の件数
    cc_count: int = 0  # Cc宛先の件数
    bcc_count: int = 0  # Bcc宛先の件数
    created_at: str

    class Config:
        from_attributes = True


class RecipientListDetailResponse(RecipientListResponse):
    members: List[RecipientMemberResponse]


class EntraUserResponse(BaseModel):
    id: str
    displayName: str
    mail: Optional[str]
    department: Optional[str]
    jobTitle: Optional[str]


class FuzzySearchResult(BaseModel):
    """ファジー検索結果"""
    email: str
    name: Optional[str]
    department: Optional[str]
    position: Optional[str] = None  # 役職
    score: float  # マッチスコア (0-1)
    match_field: str  # マッチした項目 (email / name / department / position)
    source: str  # データソース (local / entra)


class FuzzySearchResponse(BaseModel):
    """ファジー検索レスポンス"""
    results: List[FuzzySearchResult]
    total_count: int
    query: str


# Validation Schemas（Entra比較用）
class ValidationWarningDetail(BaseModel):
    """バリデーション警告の詳細"""
    field: str
    uploaded_value: str
    current_value: str


class ValidationWarning(BaseModel):
    """バリデーション警告"""
    email: str
    warning_type: str  # 'info_mismatch', 'not_found'
    message: str
    details: Optional[dict] = None


class ValidationResponse(BaseModel):
    """バリデーションレスポンス"""
    list_id: Optional[int] = None
    checked_at: str
    total_members: int
    matched: int
    mismatched: int
    not_found: int
    warnings: List[ValidationWarning]
    requires_confirmation: bool


class RecipientListDetailWithValidationResponse(RecipientListDetailResponse):
    """バリデーション結果付き宛先リスト詳細レスポンス"""
    validation_warnings: Optional[List[ValidationWarning]] = None
    requires_confirmation: bool = False


# AI Filtering Schemas（自然言語フィルタリング用）
class FilterMembersRequest(BaseModel):
    """フィルタリングリクエスト"""
    instruction: str  # 自然言語指示（例：「営業部の人だけ」）


class FilteredMemberResponse(BaseModel):
    """フィルタリング結果のメンバー"""
    id: int
    email: str
    name: Optional[str] = None
    department: Optional[str] = None
    position: Optional[str] = None
    employee_id: Optional[str] = None
    selected: bool
    reason: str

    class Config:
        from_attributes = True


class FilterMembersResponse(BaseModel):
    """フィルタリングレスポンス"""
    selected_members: List[FilteredMemberResponse]
    excluded_members: List[FilteredMemberResponse]
    summary: str
    selected_count: int
    excluded_count: int
    total_count: int


def calculate_fuzzy_score(query: str, candidate: RecipientListMember) -> tuple:
    """
    ファジー検索スコアを計算

    Returns:
        tuple: (best_score, match_field)
    """
    query_lower = query.lower()

    # メールアドレスのスコア
    email_score = 0.0
    if candidate.email:
        email_lower = candidate.email.lower()
        # 完全一致
        if email_lower == query_lower:
            email_score = 1.0
        # 部分一致（含む）
        elif query_lower in email_lower:
            email_score = 0.85
        # ファジーマッチ
        else:
            email_score = fuzz.ratio(query_lower, email_lower) / 100

    # 名前のスコア
    name_score = 0.0
    if candidate.name:
        name = candidate.name
        name_lower = name.lower()
        # 完全一致
        if name_lower == query_lower or name == query:
            name_score = 0.95
        # 部分一致（含む）
        elif query_lower in name_lower or query in name:
            name_score = 0.8
        # ファジーマッチ（日本語対応のため部分一致スコアも使用）
        else:
            # partial_ratio は部分一致に強い
            name_score = max(
                fuzz.ratio(query, name) / 100,
                fuzz.partial_ratio(query, name) / 100 * 0.9  # 部分一致は少し低く評価
            )

    # 部署のスコア
    dept_score = 0.0
    if candidate.department:
        dept = candidate.department
        dept_lower = dept.lower()
        # 完全一致
        if dept_lower == query_lower or dept == query:
            dept_score = 0.6
        # 部分一致（含む）
        elif query_lower in dept_lower or query in dept:
            dept_score = 0.55
        # ファジーマッチ
        else:
            dept_score = fuzz.ratio(query, dept) / 100 * 0.5

    # 役職のスコア
    position_score = 0.0
    if candidate.position:
        pos = candidate.position
        pos_lower = pos.lower()
        # 完全一致
        if pos_lower == query_lower or pos == query:
            position_score = 0.9
        # 部分一致（含む）
        elif query_lower in pos_lower or query in pos:
            position_score = 0.85
        # ファジーマッチ
        else:
            position_score = fuzz.ratio(query, pos) / 100 * 0.7

    # 最高スコアを採用
    scores = [
        (email_score, "email"),
        (name_score, "name"),
        (dept_score, "department"),
        (position_score, "position"),
    ]
    best_score, match_field = max(scores, key=lambda x: x[0])

    return best_score, match_field


# Endpoints
@router.get("/lists", response_model=List[RecipientListResponse])
async def list_recipient_lists(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """宛先リスト一覧を取得"""
    lists = db.query(RecipientList).filter(
        RecipientList.user_id == current_user.id
    ).order_by(RecipientList.created_at.desc()).all()

    result = []
    for lst in lists:
        # To/Cc/Bcc件数をカウント
        to_count = sum(1 for m in lst.members if (m.recipient_type or 'to') == 'to')
        cc_count = sum(1 for m in lst.members if m.recipient_type == 'cc')
        bcc_count = sum(1 for m in lst.members if m.recipient_type == 'bcc')

        result.append(RecipientListResponse(
            id=lst.id,
            name=lst.name,
            description=lst.description,
            member_count=len(lst.members),
            to_count=to_count,
            cc_count=cc_count,
            bcc_count=bcc_count,
            created_at=lst.created_at.isoformat(),
        ))

    return result


@router.post("/lists", response_model=RecipientListDetailResponse, status_code=status.HTTP_201_CREATED)
async def create_recipient_list(
    data: RecipientListCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """宛先リストを作成"""
    recipient_list = RecipientList(
        user_id=current_user.id,
        name=data.name,
        description=data.description,
    )
    db.add(recipient_list)
    db.flush()

    if data.members:
        for member in data.members:
            db.add(RecipientListMember(
                list_id=recipient_list.id,
                email=member.email,
                name=member.name,
                department=member.department,
                position=member.position,
                employee_id=member.employee_id,
                note=member.note,
                recipient_type=member.recipient_type or 'to',
            ))

    db.commit()
    db.refresh(recipient_list)

    # To/Cc/Bcc件数をカウント
    to_count = sum(1 for m in recipient_list.members if (m.recipient_type or 'to') == 'to')
    cc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'cc')
    bcc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'bcc')

    return RecipientListDetailResponse(
        id=recipient_list.id,
        name=recipient_list.name,
        description=recipient_list.description,
        member_count=len(recipient_list.members),
        to_count=to_count,
        cc_count=cc_count,
        bcc_count=bcc_count,
        created_at=recipient_list.created_at.isoformat(),
        members=[
            RecipientMemberResponse(
                id=m.id,
                email=m.email,
                name=m.name,
                department=m.department,
                position=m.position,
                employee_id=m.employee_id,
                note=m.note,
                recipient_type=m.recipient_type or 'to',
            )
            for m in recipient_list.members
        ],
    )


@router.get("/lists/{list_id}", response_model=RecipientListDetailResponse)
async def get_recipient_list(
    list_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """宛先リスト詳細を取得"""
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="Recipient list not found")

    # To/Cc/Bcc件数をカウント
    to_count = sum(1 for m in recipient_list.members if (m.recipient_type or 'to') == 'to')
    cc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'cc')
    bcc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'bcc')

    return RecipientListDetailResponse(
        id=recipient_list.id,
        name=recipient_list.name,
        description=recipient_list.description,
        member_count=len(recipient_list.members),
        to_count=to_count,
        cc_count=cc_count,
        bcc_count=bcc_count,
        created_at=recipient_list.created_at.isoformat(),
        members=[
            RecipientMemberResponse(
                id=m.id,
                email=m.email,
                name=m.name,
                department=m.department,
                position=m.position,
                employee_id=m.employee_id,
                note=m.note,
                recipient_type=m.recipient_type or 'to',
            )
            for m in recipient_list.members
        ],
    )


class RecipientListUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None


@router.put("/lists/{list_id}", response_model=RecipientListResponse)
async def update_recipient_list(
    list_id: int,
    data: RecipientListUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """宛先リストを更新"""
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    if data.name is not None:
        recipient_list.name = data.name
    if data.description is not None:
        recipient_list.description = data.description

    db.commit()
    db.refresh(recipient_list)

    # To/Cc/Bcc件数をカウント
    to_count = sum(1 for m in recipient_list.members if (m.recipient_type or 'to') == 'to')
    cc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'cc')
    bcc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'bcc')

    return RecipientListResponse(
        id=recipient_list.id,
        name=recipient_list.name,
        description=recipient_list.description,
        member_count=len(recipient_list.members),
        to_count=to_count,
        cc_count=cc_count,
        bcc_count=bcc_count,
        created_at=recipient_list.created_at.isoformat(),
    )


@router.delete("/lists/{list_id}")
async def delete_recipient_list(
    list_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """宛先リストを削除"""
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    try:
        # カスケード削除で関連メンバーも自動削除される
        db.delete(recipient_list)
        db.commit()
        return {"message": "削除しました"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"削除に失敗しました: {str(e)}")


class AddMembersRequest(BaseModel):
    """既存リストへのメンバー追加リクエスト"""
    members: List[RecipientMemberCreate]


class AddMembersResponse(BaseModel):
    """メンバー追加レスポンス"""
    added_count: int
    skipped_count: int
    message: str


@router.post("/lists/{list_id}/members", response_model=AddMembersResponse)
async def add_members_to_list(
    list_id: int,
    data: AddMembersRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """既存の宛先リストにメンバーを追加"""
    # リストを取得
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    # 既存のメールアドレスを取得（重複チェック用）
    existing_emails = {m.email.lower() for m in recipient_list.members}

    added_count = 0
    skipped_count = 0

    for member in data.members:
        if member.email.lower() in existing_emails:
            skipped_count += 1
            continue

        new_member = RecipientListMember(
            list_id=recipient_list.id,
            email=member.email,
            name=member.name,
            department=member.department,
            position=member.position,
            employee_id=member.employee_id,
            note=member.note,
            recipient_type=member.recipient_type or 'to',
        )
        db.add(new_member)
        existing_emails.add(member.email.lower())
        added_count += 1

    db.commit()

    return AddMembersResponse(
        added_count=added_count,
        skipped_count=skipped_count,
        message=f"{added_count}名を追加しました" + (f"（{skipped_count}名は既に登録済み）" if skipped_count > 0 else ""),
    )


class RecipientMemberUpdate(BaseModel):
    """メンバー更新リクエスト"""
    email: Optional[str] = None
    name: Optional[str] = None
    department: Optional[str] = None
    position: Optional[str] = None
    employee_id: Optional[str] = None
    note: Optional[str] = None
    recipient_type: Optional[str] = None


@router.put("/lists/{list_id}/members/{member_id}", response_model=RecipientMemberResponse)
async def update_member(
    list_id: int,
    member_id: int,
    data: RecipientMemberUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """既存メンバーを更新"""
    # リストを取得（権限チェック）
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    # メンバーを取得
    member = db.query(RecipientListMember).filter(
        RecipientListMember.id == member_id,
        RecipientListMember.list_id == list_id,
    ).first()

    if not member:
        raise HTTPException(status_code=404, detail="メンバーが見つかりません")

    # 更新
    if data.email is not None:
        member.email = data.email
    if data.name is not None:
        member.name = data.name if data.name else None
    if data.department is not None:
        member.department = data.department if data.department else None
    if data.position is not None:
        member.position = data.position if data.position else None
    if data.employee_id is not None:
        member.employee_id = data.employee_id if data.employee_id else None
    if data.note is not None:
        member.note = data.note if data.note else None
    if data.recipient_type is not None and data.recipient_type in ['to', 'cc', 'bcc']:
        member.recipient_type = data.recipient_type

    db.commit()
    db.refresh(member)

    return RecipientMemberResponse(
        id=member.id,
        email=member.email,
        name=member.name,
        department=member.department,
        position=member.position,
        employee_id=member.employee_id,
        note=member.note,
        recipient_type=member.recipient_type or 'to',
    )


@router.delete("/lists/{list_id}/members/{member_id}")
async def delete_member(
    list_id: int,
    member_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """メンバーを削除"""
    # リストを取得（権限チェック）
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    # メンバーを取得
    member = db.query(RecipientListMember).filter(
        RecipientListMember.id == member_id,
        RecipientListMember.list_id == list_id,
    ).first()

    if not member:
        raise HTTPException(status_code=404, detail="メンバーが見つかりません")

    db.delete(member)
    db.commit()

    return {"message": "メンバーを削除しました"}


# ==============================================================================
# Template Download Endpoints
# ==============================================================================

def create_excel_template(include_recipient_type: bool = False) -> io.BytesIO:
    """Excelテンプレートを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "宛先リスト"

    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ヘッダー定義
    if include_recipient_type:
        # 振り分け版: 宛先種別, メールアドレス, 名前, 所属, 職位, 社員番号
        headers = ["宛先種別", "メールアドレス", "名前", "所属", "職位", "社員番号"]
        col_widths = [12, 35, 20, 20, 15, 15]
    else:
        # シンプル版: メールアドレス, 名前, 所属, 職位, 社員番号
        headers = ["メールアドレス", "名前", "所属", "職位", "社員番号"]
        col_widths = [35, 20, 20, 15, 15]

    # ヘッダー行を作成
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 列幅を設定
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 振り分け版: 宛先種別列にプルダウン（データ検証）を追加
    if include_recipient_type:
        # To, Cc, Bcc のプルダウンを作成
        dv = DataValidation(
            type="list",
            formula1='"To,Cc,Bcc"',
            allow_blank=True
        )
        dv.error = "To, Cc, Bcc のいずれかを選択してください"
        dv.errorTitle = "入力エラー"
        dv.prompt = "To, Cc, Bcc から選択"
        dv.promptTitle = "宛先種別"

        # A2:A1000 にプルダウンを適用
        dv.add("A2:A1000")
        ws.add_data_validation(dv)

    # バイトストリームに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/templates/simple")
async def download_simple_template():
    """
    シンプル版テンプレートをダウンロード（全員To）

    カラム: メールアドレス, 名前, 所属, 職位, 社員番号
    """
    output = create_excel_template(include_recipient_type=False)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=recipient_template_simple.xlsx"
        }
    )


@router.get("/templates/with-types")
async def download_template_with_types():
    """
    振り分け版テンプレートをダウンロード（To/Cc/Bcc対応）

    カラム: 宛先種別(プルダウン), メールアドレス, 名前, 所属, 職位, 社員番号
    """
    output = create_excel_template(include_recipient_type=True)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=recipient_template_with_types.xlsx"
        }
    )


# ==============================================================================
# Import to Existing List Endpoint
# ==============================================================================

class ImportMembersResponse(BaseModel):
    """インポート結果レスポンス"""
    added_count: int
    skipped_count: int
    skipped_reasons: List[dict]
    message: str


@router.post("/lists/{list_id}/import", response_model=ImportMembersResponse)
async def import_members_to_list(
    list_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    既存リストにExcel/CSVファイルからメンバーをインポート

    - 既存メンバーと重複するメールアドレスはスキップ
    - 宛先種別列がある場合はTo/Cc/Bccを設定
    """
    # リストを取得
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    # ファイル形式チェック
    if not file.filename:
        raise HTTPException(status_code=400, detail="ファイル名が必要です")

    ext = file.filename.lower().split(".")[-1]
    if ext not in ["xlsx", "xls", "csv"]:
        raise HTTPException(status_code=400, detail="対応形式: xlsx, xls, csv")

    # ファイル読み込み
    content = await file.read()
    try:
        if ext == "csv":
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ファイル読み込みエラー: {str(e)}")

    # カラムマッピング
    email_col = None
    name_col = None
    dept_col = None
    position_col = None
    note_col = None
    employee_id_col = None
    recipient_type_col = None

    for col in df.columns:
        col_lower = str(col).lower()
        col_str = str(col)
        if "mail" in col_lower or "メール" in col_str or "アドレス" in col_str:
            email_col = col
        elif "名" in col_str or "name" in col_lower:
            name_col = col
        elif "部" in col_str or "所属" in col_str or "department" in col_lower:
            dept_col = col
        elif "役職" in col_str or "職位" in col_str or "position" in col_lower or "title" in col_lower or "ポジション" in col_str:
            position_col = col
        elif "備考" in col_str or "note" in col_lower or "メモ" in col_str:
            note_col = col
        elif "社員番号" in col_str or "employee" in col_lower or "番号" in col_str:
            employee_id_col = col
        elif "宛先種別" in col_str or "種別" in col_str or "type" in col_lower or "to/cc/bcc" in col_lower or "recipient_type" in col_lower:
            recipient_type_col = col

    if not email_col:
        raise HTTPException(status_code=400, detail="メールアドレス列が見つかりません")

    # 既存のメールアドレスを取得（重複チェック用）
    existing_emails = {m.email.lower() for m in recipient_list.members}

    # recipient_type 正規化関数
    def normalize_recipient_type(value) -> str:
        if not value or pd.isna(value):
            return 'to'
        value_str = str(value).strip().lower()
        if value_str in ['to', 'cc', 'bcc']:
            return value_str
        return 'to'

    # メンバー追加
    added_count = 0
    skipped_count = 0
    skipped_reasons = []

    for row_idx, row in df.iterrows():
        email = str(row[email_col]).strip() if pd.notna(row[email_col]) else ""

        # メールアドレスのバリデーション
        if not email or "@" not in email:
            skipped_count += 1
            skipped_reasons.append({
                "row": row_idx + 2,  # Excelの行番号（ヘッダー+1）
                "reason": "メールアドレスが無効です"
            })
            continue

        # 重複チェック
        if email.lower() in existing_emails:
            skipped_count += 1
            skipped_reasons.append({
                "row": row_idx + 2,
                "reason": f"既に登録されています: {email}"
            })
            continue

        # recipient_type を取得
        recipient_type = 'to'
        if recipient_type_col and recipient_type_col in row:
            recipient_type = normalize_recipient_type(row.get(recipient_type_col))

        member = RecipientListMember(
            list_id=recipient_list.id,
            email=email,
            name=str(row[name_col]).strip() if name_col and pd.notna(row.get(name_col)) else None,
            department=str(row[dept_col]).strip() if dept_col and pd.notna(row.get(dept_col)) else None,
            position=str(row[position_col]).strip() if position_col and pd.notna(row.get(position_col)) else None,
            employee_id=str(row[employee_id_col]).strip() if employee_id_col and pd.notna(row.get(employee_id_col)) else None,
            note=str(row[note_col]).strip() if note_col and pd.notna(row.get(note_col)) else None,
            recipient_type=recipient_type,
        )
        db.add(member)
        existing_emails.add(email.lower())
        added_count += 1

    if added_count == 0 and skipped_count > 0:
        raise HTTPException(status_code=400, detail="有効なデータがありません")

    db.commit()

    return ImportMembersResponse(
        added_count=added_count,
        skipped_count=skipped_count,
        skipped_reasons=skipped_reasons[:10],  # 最大10件まで
        message=f"{added_count}名を追加しました" + (f"（{skipped_count}名はスキップ）" if skipped_count > 0 else ""),
    )


@router.post("/upload", response_model=RecipientListDetailResponse)
async def upload_recipients(
    file: UploadFile = File(...),
    name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel/CSVファイルから宛先リストをアップロード"""
    # ファイル形式チェック
    if not file.filename:
        raise HTTPException(status_code=400, detail="Filename is required")

    ext = file.filename.lower().split(".")[-1]
    if ext not in ["xlsx", "xls", "csv"]:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use xlsx, xls, or csv.")

    # ファイル読み込み
    content = await file.read()
    try:
        if ext == "csv":
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")

    # カラムマッピング
    email_col = None
    name_col = None
    dept_col = None
    position_col = None
    note_col = None
    employee_id_col = None
    recipient_type_col = None

    for col in df.columns:
        col_lower = str(col).lower()
        col_str = str(col)
        if "mail" in col_lower or "メール" in col_str or "アドレス" in col_str:
            email_col = col
        elif "名" in col_str or "name" in col_lower:
            name_col = col
        elif "部" in col_str or "所属" in col_str or "department" in col_lower:
            dept_col = col
        elif "役職" in col_str or "職位" in col_str or "position" in col_lower or "title" in col_lower or "ポジション" in col_str:
            position_col = col
        elif "備考" in col_str or "note" in col_lower or "メモ" in col_str:
            note_col = col
        elif "社員番号" in col_str or "employee" in col_lower or "番号" in col_str:
            employee_id_col = col
        elif "宛先種別" in col_str or "種別" in col_str or "type" in col_lower or "to/cc/bcc" in col_lower or "recipient_type" in col_lower:
            recipient_type_col = col

    # 役職列が見つからない場合、内容から推測（「一般」「チーフ」等の値がある列）
    if not position_col:
        position_keywords = ['一般', 'チーフ', 'マネージャー', 'リーダー', '主任', '課長', '部長', '係長', '担当']
        for col in df.columns:
            if col in [email_col, name_col, dept_col, note_col]:
                continue
            sample_values = df[col].dropna().astype(str).head(20).tolist()
            if any(kw in val for val in sample_values for kw in position_keywords):
                position_col = col
                break

    if not email_col:
        raise HTTPException(status_code=400, detail="Email column not found in file")

    # 宛先リスト作成
    list_name = name or file.filename.rsplit(".", 1)[0]
    recipient_list = RecipientList(
        user_id=current_user.id,
        name=list_name,
        description=f"Uploaded from {file.filename}",
    )
    db.add(recipient_list)
    db.flush()

    # recipient_type 正規化関数
    def normalize_recipient_type(value) -> str:
        if not value or pd.isna(value):
            return 'to'
        value_str = str(value).strip().lower()
        if value_str in ['to', 'cc', 'bcc']:
            return value_str
        return 'to'

    # メンバー追加
    added = 0
    for _, row in df.iterrows():
        email = str(row[email_col]).strip() if pd.notna(row[email_col]) else ""
        if not email or "@" not in email:
            continue

        # recipient_type を取得（列がない場合や空の場合は 'to'）
        recipient_type = 'to'
        if recipient_type_col and recipient_type_col in row:
            recipient_type = normalize_recipient_type(row.get(recipient_type_col))

        member = RecipientListMember(
            list_id=recipient_list.id,
            email=email,
            name=str(row[name_col]).strip() if name_col and pd.notna(row.get(name_col)) else None,
            department=str(row[dept_col]).strip() if dept_col and pd.notna(row.get(dept_col)) else None,
            position=str(row[position_col]).strip() if position_col and pd.notna(row.get(position_col)) else None,
            employee_id=str(row[employee_id_col]).strip() if employee_id_col and pd.notna(row.get(employee_id_col)) else None,
            note=str(row[note_col]).strip() if note_col and pd.notna(row.get(note_col)) else None,
            recipient_type=recipient_type,
        )
        db.add(member)
        added += 1

    db.commit()
    db.refresh(recipient_list)

    # To/Cc/Bcc件数をカウント
    to_count = sum(1 for m in recipient_list.members if (m.recipient_type or 'to') == 'to')
    cc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'cc')
    bcc_count = sum(1 for m in recipient_list.members if m.recipient_type == 'bcc')

    return RecipientListDetailResponse(
        id=recipient_list.id,
        name=recipient_list.name,
        description=recipient_list.description,
        member_count=len(recipient_list.members),
        to_count=to_count,
        cc_count=cc_count,
        bcc_count=bcc_count,
        created_at=recipient_list.created_at.isoformat(),
        members=[
            RecipientMemberResponse(
                id=m.id,
                email=m.email,
                name=m.name,
                department=m.department,
                position=m.position,
                employee_id=m.employee_id,
                note=m.note,
                recipient_type=m.recipient_type or 'to',
            )
            for m in recipient_list.members
        ],
    )


@router.post("/lists/{list_id}/validate", response_model=ValidationResponse)
async def validate_recipient_list(
    list_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """宛先リストをEntraデータと比較してバリデーション"""
    # リストを取得
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    if not recipient_list.members:
        return ValidationResponse(
            list_id=list_id,
            checked_at=datetime.utcnow().isoformat(),
            total_members=0,
            matched=0,
            mismatched=0,
            not_found=0,
            warnings=[],
            requires_confirmation=False,
        )

    # アクセストークンを取得
    tokens = get_user_tokens(db, current_user.id)
    access_token = tokens.get("access_token") if tokens else None

    # バリデーション実行
    validation_service = EntraValidationService(db)
    result = await validation_service.validate_recipient_list(
        members=recipient_list.members,
        access_token=access_token,
    )

    return ValidationResponse(
        list_id=list_id,
        checked_at=result.checked_at.isoformat(),
        total_members=result.total,
        matched=result.matched,
        mismatched=result.mismatched,
        not_found=result.not_found,
        warnings=[
            ValidationWarning(
                email=w.email,
                warning_type=w.warning_type,
                message=w.message,
                details=w.details,
            )
            for w in result.warnings
        ],
        requires_confirmation=result.requires_confirmation,
    )


@router.post("/lists/{list_id}/filter", response_model=FilterMembersResponse)
async def filter_list_members(
    list_id: int,
    request: FilterMembersRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    自然言語指示でリストメンバーをフィルタリング

    AIを使用して、自然言語の指示に基づいてリストメンバーを選択/除外します。

    例:
    - 「営業部の人だけ選んで」
    - 「課長以上の役職の人」
    - 「山田さんを除いて」
    - 「社員番号がAで始まる人」
    """
    # リスト取得・権限チェック
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    # メンバー取得
    members = db.query(RecipientListMember).filter(
        RecipientListMember.list_id == list_id
    ).all()

    if not members:
        raise HTTPException(status_code=400, detail="リストにメンバーがいません")

    # フィルタリング指示のバリデーション
    if not request.instruction or not request.instruction.strip():
        raise HTTPException(status_code=400, detail="フィルタリング指示が空です")

    # MemberDataに変換
    member_data = [
        MemberData(
            id=m.id,
            email=m.email,
            name=m.name,
            department=m.department,
            position=m.position,
            employee_id=m.employee_id,
            note=m.note,
        )
        for m in members
    ]

    # フィルタリング実行
    try:
        filter_service = LangChainRecipientFilter()
        result = await filter_service.filter_members(member_data, request.instruction)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))

    # レスポンス構築
    return FilterMembersResponse(
        selected_members=[
            FilteredMemberResponse(
                id=m.id,
                email=m.email,
                name=m.name,
                department=m.department,
                position=m.position,
                employee_id=m.employee_id,
                selected=m.selected,
                reason=m.reason,
            )
            for m in result.selected_members
        ],
        excluded_members=[
            FilteredMemberResponse(
                id=m.id,
                email=m.email,
                name=m.name,
                department=m.department,
                position=m.position,
                employee_id=m.employee_id,
                selected=m.selected,
                reason=m.reason,
            )
            for m in result.excluded_members
        ],
        summary=result.summary,
        selected_count=len(result.selected_members),
        excluded_count=len(result.excluded_members),
        total_count=len(members),
    )


@router.get("/search", response_model=List[EntraUserResponse])
async def search_entra_users(
    q: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Entra ID（Azure AD）からユーザーを検索"""
    if len(q) < 2:
        return []

    # ユーザーのアクセストークンを取得
    tokens = get_user_tokens(db, current_user.id)
    if not tokens or not tokens.get("access_token"):
        raise HTTPException(
            status_code=403,
            detail="Entra ID access token not available. Please login with Entra ID."
        )

    access_token = tokens["access_token"]

    # Microsoft Graph API でユーザー検索
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.get(
                "https://graph.microsoft.com/v1.0/users",
                params={
                    "$filter": f"startswith(displayName,'{q}') or startswith(mail,'{q}')",
                    "$select": "id,displayName,mail,department,jobTitle",
                    "$top": "10",
                },
                headers={
                    "Authorization": f"Bearer {access_token}",
                },
            )

            if response.status_code == 401:
                raise HTTPException(
                    status_code=401,
                    detail="Access token expired. Please re-login with Entra ID."
                )

            if response.status_code != 200:
                raise HTTPException(
                    status_code=response.status_code,
                    detail=f"Graph API error: {response.text}"
                )

            data = response.json()
            users = data.get("value", [])

            return [
                EntraUserResponse(
                    id=u.get("id", ""),
                    displayName=u.get("displayName", ""),
                    mail=u.get("mail"),
                    department=u.get("department"),
                    jobTitle=u.get("jobTitle"),
                )
                for u in users
                if u.get("mail")  # メールアドレスがあるユーザーのみ
            ]

    except httpx.RequestError as e:
        raise HTTPException(status_code=500, detail=f"Failed to connect to Graph API: {str(e)}")


@router.get("/search/local", response_model=List[RecipientMemberResponse])
async def search_local_recipients(
    q: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """ローカルの宛先リストから検索"""
    if len(q) < 2:
        return []

    # ユーザーの宛先リストから検索
    members = db.query(RecipientListMember).join(RecipientList).filter(
        RecipientList.user_id == current_user.id,
    ).filter(
        (RecipientListMember.email.ilike(f"%{q}%")) |
        (RecipientListMember.name.ilike(f"%{q}%")) |
        (RecipientListMember.department.ilike(f"%{q}%"))
    ).limit(20).all()

    return [
        RecipientMemberResponse(
            id=m.id,
            email=m.email,
            name=m.name,
            department=m.department,
            position=m.position,
            note=m.note,
        )
        for m in members
    ]


@router.get("/search/fuzzy", response_model=FuzzySearchResponse)
async def fuzzy_search_recipients(
    q: str = Query(..., min_length=1, description="検索クエリ"),
    threshold: float = Query(0.4, ge=0, le=1, description="最小マッチスコア（0-1）"),
    limit: int = Query(20, ge=1, le=100, description="最大結果数"),
    include_entra: bool = Query(False, description="Entra ID検索を含めるか（デフォルト無効）"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    ファジー検索（タイプミス許容）でユーザーを検索

    - ローカル宛先リストとEntra IDの両方から検索
    - Levenshtein距離ベースのスコアリング
    - 検索対象: メールアドレス、名前、部署
    """
    results: List[FuzzySearchResult] = []
    seen_emails = set()  # 重複排除用

    # 1. ローカル宛先リストからファジー検索
    all_members = db.query(RecipientListMember).join(RecipientList).filter(
        RecipientList.user_id == current_user.id,
    ).all()

    for member in all_members:
        score, match_field = calculate_fuzzy_score(q, member)

        if score >= threshold and member.email not in seen_emails:
            results.append(FuzzySearchResult(
                email=member.email,
                name=member.name,
                department=member.department,
                score=round(score, 3),
                match_field=match_field,
                source="local",
            ))
            seen_emails.add(member.email)

    # 2. Entra ID検索（オプション、前方一致のみだがスコア計算は行う）
    if include_entra and len(q) >= 2:
        try:
            tokens = get_user_tokens(db, current_user.id)
            if tokens and tokens.get("access_token"):
                access_token = tokens["access_token"]

                async with httpx.AsyncClient(timeout=10) as client:
                    response = await client.get(
                        "https://graph.microsoft.com/v1.0/users",
                        params={
                            "$filter": f"startswith(displayName,'{q}') or startswith(mail,'{q}')",
                            "$select": "id,displayName,mail,department,jobTitle",
                            "$top": "20",
                        },
                        headers={
                            "Authorization": f"Bearer {access_token}",
                        },
                    )

                    if response.status_code == 200:
                        data = response.json()
                        users = data.get("value", [])

                        for u in users:
                            email = u.get("mail")
                            if not email or email in seen_emails:
                                continue

                            # 仮のRecipientListMemberとしてスコア計算
                            class TempMember:
                                def __init__(self, email, name, department, position=None):
                                    self.email = email
                                    self.name = name
                                    self.department = department
                                    self.position = position

                            temp = TempMember(
                                email=email,
                                name=u.get("displayName"),
                                department=u.get("department"),
                                position=u.get("jobTitle"),
                            )
                            score, match_field = calculate_fuzzy_score(q, temp)

                            if score >= threshold:
                                results.append(FuzzySearchResult(
                                    email=email,
                                    name=u.get("displayName"),
                                    department=u.get("department"),
                                    score=round(score, 3),
                                    match_field=match_field,
                                    source="entra",
                                ))
                                seen_emails.add(email)

        except Exception as e:
            # Entra ID検索に失敗してもローカル結果は返す
            print(f"Entra ID search failed: {e}")

    # 3. スコア順にソート
    results.sort(key=lambda x: x.score, reverse=True)

    # 4. 上限適用
    results = results[:limit]

    return FuzzySearchResponse(
        results=results,
        total_count=len(results),
        query=q,
    )


@router.get("/search/unified", response_model=FuzzySearchResponse)
async def unified_search_recipients(
    q: str = Query(..., min_length=1, description="検索クエリ"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    統合検索（フロントエンド向け簡易API）

    - ファジー検索をデフォルト設定で実行
    - To/Cc/Bcc入力欄からの検索に最適化
    """
    return await fuzzy_search_recipients(
        q=q,
        threshold=0.3,  # 緩めの閾値で多くの候補を返す
        limit=15,
        include_entra=False,  # Entra検索は裏側でのみ使用（UI非表示）
        db=db,
        current_user=current_user,
    )


# ==============================================================================
# AI Agent for Natural Language Recipient Search
# ==============================================================================

class AIAgentMessage(BaseModel):
    """AIエージェントチャットメッセージ"""
    role: str  # user / assistant
    content: str


class AIAgentChatRequest(BaseModel):
    """AIエージェントチャットリクエスト"""
    messages: List[AIAgentMessage]


class AIAgentRecipient(BaseModel):
    """AIエージェントが見つけた宛先"""
    email: str
    name: Optional[str]
    department: Optional[str]
    position: Optional[str] = None  # 役職
    reason: str  # なぜこの人が選ばれたか


class AIAgentChatResponse(BaseModel):
    """AIエージェントチャットレスポンス"""
    message: str
    recipients: Optional[List[AIAgentRecipient]] = None
    action: Optional[str] = None  # asking / suggesting / confirming / done


def build_recipient_context(members: List[RecipientListMember], lists: List[RecipientList]) -> str:
    """宛先リストのコンテキストを構築（メールアドレス・役職付き）"""
    context_parts = []

    # リスト情報
    context_parts.append("【利用可能な宛先リスト】")
    for lst in lists:
        list_members = [m for m in members if m.list_id == lst.id]
        context_parts.append(f"- {lst.name}（{len(list_members)}名）")

    context_parts.append("\n【登録されている宛先一覧】")
    context_parts.append("※形式: 名前 <メールアドレス> [役職]")

    # 役職でグループ化（役職がある場合）
    positions = set(m.position for m in members if m.position)
    if positions:
        context_parts.append("\n【役職一覧】")
        for pos in sorted(positions):
            count = len([m for m in members if m.position == pos])
            context_parts.append(f"  - {pos}（{count}名）")

    # 部署別にグループ化
    by_dept = {}
    for m in members:
        dept = m.department or "部署未設定"
        if dept not in by_dept:
            by_dept[dept] = []
        by_dept[dept].append(m)

    for dept, dept_members in sorted(by_dept.items()):
        context_parts.append(f"\n■ {dept}")
        for m in dept_members:
            name = m.name or "(名前なし)"
            pos_info = f" [{m.position}]" if m.position else ""
            context_parts.append(f"  - {name} <{m.email}>{pos_info}")

    return "\n".join(context_parts)


def parse_ai_search_intent(ai_response: str, all_members: List[RecipientListMember]) -> List[AIAgentRecipient]:
    """
    AIの応答から検索意図を解析し、該当する宛先を抽出

    AIの応答に含まれる部署名や条件に基づいて宛先をフィルタリング
    """
    results = []
    response_lower = ai_response.lower()

    for member in all_members:
        reason = None

        # 部署名が応答に含まれているか
        if member.department:
            dept_lower = member.department.lower()
            # 部署名の一部が応答に含まれていれば該当
            for dept_part in member.department.split():
                if dept_part in ai_response or dept_part.lower() in response_lower:
                    reason = f"{member.department}所属"
                    break

        # 名前が応答に含まれているか
        if not reason and member.name:
            if member.name in ai_response:
                reason = "名前が一致"

        if reason:
            results.append(AIAgentRecipient(
                email=member.email,
                name=member.name,
                department=member.department,
                reason=reason,
            ))

    return results


@router.post("/ai-agent/chat", response_model=AIAgentChatResponse)
async def ai_agent_chat(
    data: AIAgentChatRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    AIエージェントとのチャット（構造化出力対応）

    自然言語で宛先を検索・選択できます。
    例：
    - 「営業部の人に送りたい」
    - 「田中さんと鈴木さんに送りたい」
    - 「全員」
    """
    if not settings.openai_api_key:
        raise HTTPException(
            status_code=503,
            detail="OpenAI API key is not configured"
        )

    # ユーザーの宛先リストとメンバーを取得
    lists = db.query(RecipientList).filter(
        RecipientList.user_id == current_user.id
    ).all()

    all_members = db.query(RecipientListMember).join(RecipientList).filter(
        RecipientList.user_id == current_user.id
    ).all()

    # 宛先のコンテキストを構築
    recipient_context = build_recipient_context(all_members, lists)

    # メールアドレス→メンバーのマッピング
    email_to_member = {m.email.lower(): m for m in all_members}

    # システムプロンプト（JSON出力指示付き）
    system_prompt = f"""あなたはメール宛先選択を支援するAIアシスタントです。
ユーザーが自然言語で宛先を指定するのを手助けしてください。

{recipient_context}

【重要な応答ルール】
必ず以下のJSON形式で応答してください：
{{
  "message": "ユーザーへのメッセージ（日本語）",
  "selected_emails": ["email1@example.com", "email2@example.com"],
  "action": "suggesting" または "asking" または "done"
}}

- selected_emails: 選択した宛先のメールアドレスの配列。宛先を提案する場合は必ずメールアドレスを入れる。質問中や該当なしの場合は空配列[]
- action:
  - "suggesting": 宛先を提案している
  - "asking": 確認の質問をしている
  - "done": 処理完了

【判断ルール】
- 「全員」「みんな」「全部」→ リスト内の全メールアドレスを selected_emails に入れる
- 部署名が指定された → その部署の全員のメールアドレスを selected_emails に入れる
- 役職名が指定された（チーフ、マネージャー、主任等）→ その役職の全員のメールアドレスを selected_emails に入れる
- 人名が指定された → その人のメールアドレスを selected_emails に入れる
- 「〜以外」「〜を除く」「〜抜き」→ 指定された人/部署/役職を【除外】し、それ以外の全員を selected_emails に入れる
  - 例: 「田中さん以外」→ 田中さん以外の全員を返す
  - 例: 「営業部以外」→ 営業部以外の全員を返す
  - 例: 「チーフ以外」→ チーフ以外の全員を返す
- 曖昧な場合 → selected_emails を空にして action を "asking" にし、質問する

例1:
ユーザー: 営業部の人
応答: {{"message": "営業部の2名を宛先に追加します。", "selected_emails": ["tanaka@example.com", "suzuki@example.com"], "action": "suggesting"}}

例2:
ユーザー: 田中さんだけ
応答: {{"message": "田中さんを宛先に追加します。", "selected_emails": ["tanaka@example.com"], "action": "suggesting"}}

例3:
ユーザー: 全員
応答: {{"message": "全員（5名）を宛先に追加します。", "selected_emails": ["a@ex.com", "b@ex.com", ...], "action": "suggesting"}}

例4（除外パターン）:
ユーザー: 田中さん以外
応答: {{"message": "田中さんを除く4名を宛先に追加します。", "selected_emails": ["suzuki@ex.com", "yamada@ex.com", ...], "action": "suggesting"}}

例5（役職検索）:
ユーザー: チーフ
応答: {{"message": "チーフの3名を宛先に追加します。", "selected_emails": ["chief1@ex.com", "chief2@ex.com", "chief3@ex.com"], "action": "suggesting"}}"""

    # メッセージ構築
    messages = [{"role": "system", "content": system_prompt}]
    for msg in data.messages:
        messages.append({"role": msg.role, "content": msg.content})

    try:
        client = openai.OpenAI(api_key=settings.openai_api_key)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            temperature=0.3,  # より確定的な応答
            max_tokens=1000,
            response_format={"type": "json_object"},  # JSON モード
        )

        ai_response_text = response.choices[0].message.content

        # JSON パース
        try:
            ai_data = json.loads(ai_response_text)
        except json.JSONDecodeError:
            # JSONパース失敗時はフォールバック
            return AIAgentChatResponse(
                message="申し訳ありません、もう一度お試しください。",
                recipients=None,
                action="asking",
            )

        ai_message = ai_data.get("message", "")
        selected_emails = ai_data.get("selected_emails", [])
        action = ai_data.get("action", "asking")

        # メールアドレスからメンバー情報を取得
        suggested_recipients = []
        for email in selected_emails:
            email_lower = email.lower()
            if email_lower in email_to_member:
                member = email_to_member[email_lower]
                reason = member.position or member.department or "選択済み"
                suggested_recipients.append(AIAgentRecipient(
                    email=member.email,
                    name=member.name,
                    department=member.department,
                    position=member.position,
                    reason=reason,
                ))
            else:
                # ファジーマッチでフォールバック
                for m_email, member in email_to_member.items():
                    if fuzz.ratio(email_lower, m_email) > 85:
                        reason = member.position or member.department or "選択済み"
                        suggested_recipients.append(AIAgentRecipient(
                            email=member.email,
                            name=member.name,
                            department=member.department,
                            position=member.position,
                            reason=reason,
                        ))
                        break

        return AIAgentChatResponse(
            message=ai_message,
            recipients=suggested_recipients if suggested_recipients else None,
            action=action,
        )

    except openai.APIError as e:
        raise HTTPException(status_code=503, detail=f"OpenAI API error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI agent chat failed: {str(e)}")


class AIAgentChatWithListRequest(BaseModel):
    """特定リスト内でのAIエージェントチャットリクエスト"""
    list_id: int
    messages: List[AIAgentMessage]


@router.post("/ai-agent/chat-with-list", response_model=AIAgentChatResponse)
async def ai_agent_chat_with_list(
    data: AIAgentChatWithListRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    特定の宛先リスト内でのAIエージェントチャット

    選択したリスト内のメンバーのみを対象に、自然言語で検索・選択できます。
    どんな自然言語の指示にも対応します。

    例：
    - 「営業部の人」
    - 「田中さんと鈴木さん」
    - 「チーフ全員」
    - 「マネージャー以外」
    - 「一番上の3人」
    - 「メールアドレスに sales が含まれる人」
    """
    if not settings.openai_api_key:
        raise HTTPException(
            status_code=503,
            detail="OpenAI API key is not configured"
        )

    # 指定されたリストを取得
    recipient_list = db.query(RecipientList).filter(
        RecipientList.id == data.list_id,
        RecipientList.user_id == current_user.id,
    ).first()

    if not recipient_list:
        raise HTTPException(status_code=404, detail="リストが見つかりません")

    # リスト内のメンバーを取得
    members = db.query(RecipientListMember).filter(
        RecipientListMember.list_id == data.list_id
    ).all()

    if not members:
        return AIAgentChatResponse(
            message="このリストにはメンバーが登録されていません。",
            recipients=None,
            action="done",
        )

    # メンバー情報をコンテキストとして構築
    member_context_parts = [f"【リスト名】{recipient_list.name}（{len(members)}名）\n"]

    # 役職でグループ化（役職がある場合）
    positions = set(m.position for m in members if m.position)
    if positions:
        member_context_parts.append("【役職一覧】")
        for pos in sorted(positions):
            count = len([m for m in members if m.position == pos])
            member_context_parts.append(f"  - {pos}（{count}名）")

    # 部署でグループ化
    by_dept = {}
    for m in members:
        dept = m.department or "部署未設定"
        if dept not in by_dept:
            by_dept[dept] = []
        by_dept[dept].append(m)

    member_context_parts.append("\n【メンバー一覧】")
    for dept, dept_members in sorted(by_dept.items()):
        member_context_parts.append(f"\n■ {dept}")
        for m in dept_members:
            name = m.name or "(名前なし)"
            pos_info = f" [{m.position}]" if m.position else ""
            member_context_parts.append(f"  - {name} <{m.email}>{pos_info}")

    member_context = "\n".join(member_context_parts)

    # メールアドレス→メンバーのマッピング
    email_to_member = {m.email.lower(): m for m in members}

    # システムプロンプト
    system_prompt = f"""あなたはメール宛先選択を支援するAIアシスタントです。
ユーザーが自然言語で宛先を指定するのを手助けしてください。

{member_context}

【重要な応答ルール】
必ず以下のJSON形式で応答してください：
{{
  "message": "ユーザーへのメッセージ（日本語）",
  "selected_emails": ["email1@example.com", "email2@example.com"],
  "action": "suggesting" または "asking" または "done"
}}

- selected_emails: 選択した宛先のメールアドレスの配列。宛先を提案する場合は必ず上記リストにあるメールアドレスを入れる。質問中や該当なしの場合は空配列[]
- action:
  - "suggesting": 宛先を提案している
  - "asking": 確認の質問をしている
  - "done": 処理完了

【判断ルール - あらゆる自然言語に対応】
- 「全員」「みんな」「全部」→ リスト内の全メールアドレスを返す
- 部署名が指定 → その部署の全員を返す
- 役職名が指定（チーフ、マネージャー、主任等）→ その役職の全員を返す
- 人名が指定 → その人を返す（部分一致でも可）
- 「〜以外」「〜を除く」「〜抜き」→ 指定された条件を【除外】し、残りを返す
- 複数条件「AとB」「AやB」→ 両方を返す
- 数量指定「最初の3人」「上から5人」→ 該当数だけ返す
- 属性検索「メールにsalesが含まれる」→ 条件に合う人を返す
- 曖昧な場合 → 質問して確認する

【重要】
- 上記リストに存在するメールアドレスのみを返してください
- 存在しないメールアドレスを生成しないでください"""

    # メッセージ構築
    messages = [{"role": "system", "content": system_prompt}]
    for msg in data.messages:
        messages.append({"role": msg.role, "content": msg.content})

    try:
        client = openai.OpenAI(api_key=settings.openai_api_key)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            temperature=0.3,
            max_tokens=1000,
            response_format={"type": "json_object"},
        )

        ai_response_text = response.choices[0].message.content

        # JSON パース
        try:
            ai_data = json.loads(ai_response_text)
        except json.JSONDecodeError:
            return AIAgentChatResponse(
                message="申し訳ありません、もう一度お試しください。",
                recipients=None,
                action="asking",
            )

        ai_message = ai_data.get("message", "")
        selected_emails = ai_data.get("selected_emails", [])
        action = ai_data.get("action", "asking")

        # メールアドレスからメンバー情報を取得（リスト内のみ）
        suggested_recipients = []
        for email in selected_emails:
            email_lower = email.lower()
            if email_lower in email_to_member:
                member = email_to_member[email_lower]
                reason = member.position or member.department or "選択済み"
                suggested_recipients.append(AIAgentRecipient(
                    email=member.email,
                    name=member.name,
                    department=member.department,
                    position=member.position,
                    reason=reason,
                ))
            else:
                # ファジーマッチでフォールバック（リスト内のみ）
                for m_email, member in email_to_member.items():
                    if fuzz.ratio(email_lower, m_email) > 85:
                        reason = member.position or member.department or "選択済み"
                        suggested_recipients.append(AIAgentRecipient(
                            email=member.email,
                            name=member.name,
                            department=member.department,
                            position=member.position,
                            reason=reason,
                        ))
                        break

        return AIAgentChatResponse(
            message=ai_message,
            recipients=suggested_recipients if suggested_recipients else None,
            action=action,
        )

    except openai.APIError as e:
        raise HTTPException(status_code=503, detail=f"OpenAI API error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI agent chat failed: {str(e)}")


@router.post("/ai-agent/search", response_model=AIAgentChatResponse)
async def ai_agent_search(
    query: str = Query(..., min_length=1, description="自然言語クエリ"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    自然言語での宛先検索（単発クエリ）

    チャットではなく、一発で検索結果を返します。
    例：「営業部」「田中さん」「マーケティングチーム」
    """
    # ユーザーの宛先メンバーを取得
    all_members = db.query(RecipientListMember).join(RecipientList).filter(
        RecipientList.user_id == current_user.id
    ).all()

    results = []

    # 部署検索
    for member in all_members:
        match_reason = None

        # 部署マッチ
        if member.department and query in member.department:
            match_reason = f"{member.department}所属"

        # 名前マッチ
        elif member.name and query in member.name:
            match_reason = "名前が一致"

        # メールマッチ
        elif query.lower() in member.email.lower():
            match_reason = "メールアドレスが一致"

        if match_reason:
            results.append(AIAgentRecipient(
                email=member.email,
                name=member.name,
                department=member.department,
                reason=match_reason,
            ))

    if results:
        return AIAgentChatResponse(
            message=f"「{query}」で{len(results)}件の宛先が見つかりました。",
            recipients=results,
            action="suggesting",
        )
    else:
        return AIAgentChatResponse(
            message=f"「{query}」に該当する宛先が見つかりませんでした。別のキーワードで検索してみてください。",
            recipients=None,
            action="asking",
        )
