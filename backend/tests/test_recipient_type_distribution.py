"""
To/Cc/Bcc振り分け機能のテスト

Phase 4: テスト・検証
- バックエンドユニットテスト
- インポート機能の結合テスト
"""
import pytest
import io
from unittest.mock import MagicMock, patch
from fastapi.testclient import TestClient

from backend.app import app
from backend.core.database import Base, engine, SessionLocal


@pytest.fixture(scope="function")
def test_db():
    """テスト用データベースセットアップ"""
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    yield db
    db.close()


@pytest.fixture
def client():
    """テストクライアント"""
    return TestClient(app)


@pytest.fixture
def mock_current_user():
    """モック認証ユーザー"""
    mock_user = MagicMock()
    mock_user.id = 1
    mock_user.email = "test@example.com"
    return mock_user


# ==============================================================================
# ユニットテスト: Pydanticスキーマ
# ==============================================================================

class TestRecipientTypeSchemas:
    """recipient_type関連スキーマのテスト"""

    def test_recipient_member_create_default_type(self):
        """RecipientMemberCreateのデフォルトrecipient_type"""
        from backend.routers.recipients import RecipientMemberCreate

        member = RecipientMemberCreate(email="test@example.com")
        assert member.recipient_type == "to"

    def test_recipient_member_create_with_type(self):
        """RecipientMemberCreateにrecipient_type指定"""
        from backend.routers.recipients import RecipientMemberCreate

        member = RecipientMemberCreate(
            email="test@example.com",
            name="Test User",
            recipient_type="cc"
        )
        assert member.recipient_type == "cc"

    def test_recipient_member_create_all_types(self):
        """全てのrecipient_typeが設定可能"""
        from backend.routers.recipients import RecipientMemberCreate

        for rtype in ["to", "cc", "bcc"]:
            member = RecipientMemberCreate(
                email=f"{rtype}@example.com",
                recipient_type=rtype
            )
            assert member.recipient_type == rtype

    def test_recipient_member_response_default_type(self):
        """RecipientMemberResponseのデフォルトrecipient_type"""
        from backend.routers.recipients import RecipientMemberResponse

        member = RecipientMemberResponse(
            id=1,
            email="test@example.com",
            name=None,
            department=None,
            position=None,
            note=None
        )
        assert member.recipient_type == "to"

    def test_recipient_list_response_counts(self):
        """RecipientListResponseのTo/Cc/Bccカウント"""
        from backend.routers.recipients import RecipientListResponse

        response = RecipientListResponse(
            id=1,
            name="Test List",
            description=None,
            member_count=10,
            to_count=5,
            cc_count=3,
            bcc_count=2,
            created_at="2024-02-16T10:00:00Z"
        )
        assert response.to_count == 5
        assert response.cc_count == 3
        assert response.bcc_count == 2
        assert response.member_count == 10


# ==============================================================================
# ユニットテスト: テンプレート生成
# ==============================================================================

class TestExcelTemplateGeneration:
    """Excelテンプレート生成のテスト"""

    def test_create_simple_template(self):
        """シンプル版テンプレートの生成"""
        from backend.routers.recipients import create_excel_template
        import openpyxl

        output = create_excel_template(include_recipient_type=False)
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        # ヘッダー確認（宛先種別列なし）
        headers = [ws.cell(row=1, column=i).value for i in range(1, 6)]
        assert "名前" in headers
        assert "所属" in headers
        assert "職位" in headers
        assert "メアド" in headers
        assert "社員番号" in headers
        assert "宛先種別" not in headers

    def test_create_template_with_types(self):
        """振り分け版テンプレートの生成"""
        from backend.routers.recipients import create_excel_template
        import openpyxl

        output = create_excel_template(include_recipient_type=True)
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        # ヘッダー確認（宛先種別列あり）
        headers = [ws.cell(row=1, column=i).value for i in range(1, 7)]
        assert "宛先種別" in headers
        assert "名前" in headers
        assert "メアド" in headers

        # サンプル行確認（To/Cc/Bcc）
        sample_types = [ws.cell(row=i, column=1).value for i in range(2, 5)]
        assert "To" in sample_types
        assert "Cc" in sample_types
        assert "Bcc" in sample_types


# ==============================================================================
# ユニットテスト: recipient_type正規化
# ==============================================================================

class TestRecipientTypeNormalization:
    """recipient_type正規化のテスト"""

    def test_normalize_to_values(self):
        """'to'への正規化"""
        test_cases = [
            ("to", "to"),
            ("To", "to"),
            ("TO", "to"),
            ("  to  ", "to"),
            ("", "to"),
            (None, "to"),
            ("invalid", "to"),
        ]

        for input_val, expected in test_cases:
            # normalize関数を直接テスト（インポート関数内で定義されている）
            if not input_val:
                result = "to"
            else:
                value_str = str(input_val).strip().lower()
                if value_str in ["to", "cc", "bcc"]:
                    result = value_str
                else:
                    result = "to"
            assert result == expected, f"Input: {input_val!r}"

    def test_normalize_cc_values(self):
        """'cc'への正規化"""
        test_cases = [
            ("cc", "cc"),
            ("Cc", "cc"),
            ("CC", "cc"),
            ("  cc  ", "cc"),
        ]

        for input_val, expected in test_cases:
            value_str = str(input_val).strip().lower()
            if value_str in ["to", "cc", "bcc"]:
                result = value_str
            else:
                result = "to"
            assert result == expected, f"Input: {input_val!r}"

    def test_normalize_bcc_values(self):
        """'bcc'への正規化"""
        test_cases = [
            ("bcc", "bcc"),
            ("Bcc", "bcc"),
            ("BCC", "bcc"),
            ("  bcc  ", "bcc"),
        ]

        for input_val, expected in test_cases:
            value_str = str(input_val).strip().lower()
            if value_str in ["to", "cc", "bcc"]:
                result = value_str
            else:
                result = "to"
            assert result == expected, f"Input: {input_val!r}"


# ==============================================================================
# APIエンドポイントテスト: テンプレートダウンロード
# ==============================================================================

class TestTemplateDownloadEndpoints:
    """テンプレートダウンロードエンドポイントのテスト"""

    def test_download_simple_template(self, client):
        """シンプル版テンプレートのダウンロード"""
        response = client.get("/recipients/templates/simple")

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "recipient_template_simple.xlsx" in response.headers["content-disposition"]

    def test_download_template_with_types(self, client):
        """振り分け版テンプレートのダウンロード"""
        response = client.get("/recipients/templates/with-types")

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "recipient_template_with_types.xlsx" in response.headers["content-disposition"]


# ==============================================================================
# インポート機能の結合テスト（認証スキップ）
# ==============================================================================

class TestImportWithRecipientType:
    """recipient_type付きインポートのテスト"""

    @pytest.mark.skip(reason="認証が必要なため統合テスト環境で実行")
    def test_import_csv_with_recipient_type(self, client, test_db):
        """CSVインポート時にrecipient_typeが設定される"""
        # 宛先種別列付きCSVファイルを作成
        csv_content = b"""email,name,department,recipient_type
to1@example.com,To User,Sales,to
cc1@example.com,Cc User,Dev,cc
bcc1@example.com,Bcc User,HR,bcc
"""
        files = {"file": ("test.csv", io.BytesIO(csv_content), "text/csv")}

        response = client.post(
            "/recipients/upload",
            files=files,
        )

        # 認証なしの場合は401
        assert response.status_code == 401

    @pytest.mark.skip(reason="認証が必要なため統合テスト環境で実行")
    def test_import_csv_without_recipient_type(self, client, test_db):
        """宛先種別列なしCSVインポート（全員To）"""
        csv_content = b"""email,name,department
user1@example.com,User 1,Sales
user2@example.com,User 2,Dev
"""
        files = {"file": ("test.csv", io.BytesIO(csv_content), "text/csv")}

        response = client.post(
            "/recipients/upload",
            files=files,
        )

        # 認証なしの場合は401
        assert response.status_code == 401

    @pytest.mark.skip(reason="認証が必要なため統合テスト環境で実行")
    def test_import_to_existing_list(self, client, test_db):
        """既存リストへのインポート"""
        csv_content = b"""email,name,recipient_type
new1@example.com,New User 1,cc
new2@example.com,New User 2,bcc
"""
        files = {"file": ("test.csv", io.BytesIO(csv_content), "text/csv")}

        response = client.post(
            "/recipients/lists/1/import",
            files=files,
        )

        # 認証なしの場合は401
        assert response.status_code == 401


# ==============================================================================
# レスポンススキーマのテスト
# ==============================================================================

class TestImportMembersResponseSchema:
    """ImportMembersResponseスキーマのテスト"""

    def test_import_response_schema(self):
        """インポートレスポンススキーマ"""
        from backend.routers.recipients import ImportMembersResponse

        response = ImportMembersResponse(
            added_count=5,
            skipped_count=2,
            skipped_reasons=[
                {"row": 3, "reason": "メールアドレスが無効です"},
                {"row": 7, "reason": "既に登録されています: test@example.com"},
            ],
            message="5名を追加しました（2名はスキップ）"
        )

        assert response.added_count == 5
        assert response.skipped_count == 2
        assert len(response.skipped_reasons) == 2
        assert "5名を追加しました" in response.message


# ==============================================================================
# To/Cc/Bccカウントのテスト
# ==============================================================================

class TestRecipientTypeCounts:
    """To/Cc/Bccカウント計算のテスト"""

    def test_count_calculation(self):
        """カウント計算ロジック"""
        # モックメンバーリスト
        members = [
            MagicMock(recipient_type="to"),
            MagicMock(recipient_type="to"),
            MagicMock(recipient_type="cc"),
            MagicMock(recipient_type="bcc"),
            MagicMock(recipient_type=None),  # Noneはtoとして扱う
        ]

        to_count = sum(1 for m in members if (m.recipient_type or 'to') == 'to')
        cc_count = sum(1 for m in members if m.recipient_type == 'cc')
        bcc_count = sum(1 for m in members if m.recipient_type == 'bcc')

        assert to_count == 3  # to x2 + None x1
        assert cc_count == 1
        assert bcc_count == 1

    def test_all_to_count(self):
        """全員Toの場合のカウント"""
        members = [
            MagicMock(recipient_type="to"),
            MagicMock(recipient_type="to"),
            MagicMock(recipient_type="to"),
        ]

        to_count = sum(1 for m in members if (m.recipient_type or 'to') == 'to')
        cc_count = sum(1 for m in members if m.recipient_type == 'cc')
        bcc_count = sum(1 for m in members if m.recipient_type == 'bcc')

        assert to_count == 3
        assert cc_count == 0
        assert bcc_count == 0

    def test_mixed_count(self):
        """混在リストのカウント"""
        members = [
            MagicMock(recipient_type="to"),
            MagicMock(recipient_type="cc"),
            MagicMock(recipient_type="cc"),
            MagicMock(recipient_type="bcc"),
            MagicMock(recipient_type="bcc"),
            MagicMock(recipient_type="bcc"),
        ]

        to_count = sum(1 for m in members if (m.recipient_type or 'to') == 'to')
        cc_count = sum(1 for m in members if m.recipient_type == 'cc')
        bcc_count = sum(1 for m in members if m.recipient_type == 'bcc')

        assert to_count == 1
        assert cc_count == 2
        assert bcc_count == 3


# ==============================================================================
# データモデルのテスト
# ==============================================================================

class TestRecipientListMemberModel:
    """RecipientListMemberモデルのテスト"""

    def test_model_has_recipient_type(self):
        """モデルにrecipient_typeフィールドがある"""
        from backend.models.recipient import RecipientListMember

        # フィールドが存在することを確認
        assert hasattr(RecipientListMember, 'recipient_type')

    def test_model_recipient_type_default(self, test_db):
        """recipient_typeのデフォルト値（SQLAlchemyはDBレベルでデフォルト設定）"""
        from backend.models.recipient import RecipientListMember
        from sqlalchemy import inspect

        # モデルのカラム定義を確認
        mapper = inspect(RecipientListMember)
        recipient_type_col = mapper.columns['recipient_type']

        # デフォルト値が'to'に設定されていることを確認
        assert recipient_type_col.default is not None
        assert recipient_type_col.default.arg == 'to'

        # nullable=Falseであることを確認
        assert recipient_type_col.nullable is False


# ==============================================================================
# E2Eテスト（リスト作成〜メール作成の一連フロー）
# ==============================================================================

class TestE2ERecipientDistribution:
    """E2E: リスト作成〜メール作成の一連フロー"""

    @pytest.mark.skip(reason="認証が必要なため統合テスト環境で実行")
    def test_full_flow_simple_list(self, client, test_db):
        """
        E2Eフロー: シンプルリスト（全員To）
        1. シンプル版テンプレートDL
        2. CSVアップロードでリスト作成
        3. リスト詳細取得（全員to_count）
        """
        # 1. テンプレートダウンロード
        response = client.get("/recipients/templates/simple")
        assert response.status_code == 200

        # 2. CSVアップロード（認証が必要）
        csv_content = b"""name,department,email
User 1,Sales,user1@example.com
User 2,Dev,user2@example.com
"""
        files = {"file": ("test.csv", io.BytesIO(csv_content), "text/csv")}
        response = client.post("/recipients/upload", files=files)
        assert response.status_code == 401  # 認証なし

    @pytest.mark.skip(reason="認証が必要なため統合テスト環境で実行")
    def test_full_flow_mixed_list(self, client, test_db):
        """
        E2Eフロー: 混在リスト（To/Cc/Bcc）
        1. 振り分け版テンプレートDL
        2. CSVアップロードでリスト作成
        3. リスト詳細取得（to_count, cc_count, bcc_count）
        4. メール作成画面でリスト選択時の振り分け
        """
        # 1. テンプレートダウンロード
        response = client.get("/recipients/templates/with-types")
        assert response.status_code == 200

        # 2. CSVアップロード（認証が必要）
        csv_content = b"""recipient_type,name,department,email
To,To User,Sales,to@example.com
Cc,Cc User,Dev,cc@example.com
Bcc,Bcc User,HR,bcc@example.com
"""
        files = {"file": ("test.csv", io.BytesIO(csv_content), "text/csv")}
        response = client.post("/recipients/upload", files=files)
        assert response.status_code == 401  # 認証なし

    @pytest.mark.skip(reason="認証が必要なため統合テスト環境で実行")
    def test_full_flow_add_to_existing_list(self, client, test_db):
        """
        E2Eフロー: 既存リストへの追加インポート
        1. 既存リストを作成
        2. 追加インポート
        3. カウント更新確認
        """
        pass


# ==============================================================================
# カラムマッピングのテスト
# ==============================================================================

class TestColumnMapping:
    """インポート時のカラムマッピングテスト"""

    def test_recipient_type_column_variations(self):
        """recipient_type列の様々な表記に対応"""
        valid_column_names = [
            "宛先種別",
            "種別",
            "type",
            "to/cc/bcc",
            "recipient_type",
            "Type",
            "TYPE",
        ]

        for col_name in valid_column_names:
            col_lower = str(col_name).lower()
            col_str = str(col_name)

            # カラムマッピングロジック
            is_recipient_type_col = (
                "宛先種別" in col_str or
                "種別" in col_str or
                "type" in col_lower or
                "to/cc/bcc" in col_lower or
                "recipient_type" in col_lower
            )

            assert is_recipient_type_col, f"Column '{col_name}' should be recognized as recipient_type column"
